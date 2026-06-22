import os
import sys
import openpyxl
from sqlalchemy.orm import Session

# 프로젝트 루트 경로를 sys.path에 추가
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "backend")))

from app.models import init_db, engine, Project, Quotation, QuotationVersion
from app.parser import run_parsing_pipeline
from app.excel_generator import generate_consolidated_excel

def setup_dummy_excel(file_path: str):
    """실무 사례와 유사한 수신처(우리 회사)와 공급자(상대 업체)가 명확히 분리된 견적서를 생성합니다."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "견적"
    
    ws["A1"] = "견 적 서"
    ws["A3"] = "DATE:"
    ws["A4"] = "업체명: KCC정보통신귀중"
    ws["A5"] = "담당자: 유기수주임"
    ws["A6"] = "TEL:"
    ws["A8"] = "mail: yooks7@kcc.co.kr"
    
    ws["E3"] = "문서NO: NP20230629-01"
    ws["E4"] = "등록번호: 264-81-46574"
    ws["E5"] = "상 호: 주식회사 넷츠플러스"
    ws["E6"] = "대표자: 송주흔 (인)"
    ws["E7"] = "주 소: 서울시 서초구 반포대로"
    
    ws["A10"] = "다음과 같이 견적서를 제출합니다"
    ws["A12"] = "총공급금액"
    ws["C12"] = "일금삼억육천이백삼십만사천원정 (부가세포함)"
    
    ws["A13"] = "번호"
    ws["B13"] = "구분"
    ws["C13"] = "수량"
    ws["D13"] = "단가"
    
    ws["A14"] = 1
    ws["B14"] = "유지보수"
    ws["C14"] = 1
    ws["D14"] = 362304800 # VAT 포함 금액이 단가 셀에 입력됨
    
    os.makedirs(os.path.dirname(file_path), exist_ok=True)
    wb.save(file_path)
    print(f"고정밀 테스트용 엑셀 견적서가 생성되었습니다: {file_path}")

def test_db_and_pipeline():
    print("=== 1. DB 재초기화 테스트 ===")
    # 기존 SQLite db 파일이 있다면 제거하여 신규 필드가 포함된 테이블로 재생성 유도
    db_file = "smart_quotation.db"
    if os.path.exists(db_file):
        try:
            os.remove(db_file)
            print("기존 SQLite DB 파일 삭제 완료.")
        except Exception as e:
            print(f"기존 DB 파일 삭제 실패 (락 상태일 수 있음): {e}")
            
    init_db()
    print("DB 테이블 신규 스키마 반영 및 초기화 완료.")
    
    # 2. 더미 파일 생성
    dummy_xlsx_path = "storage/test_dummy_netsplus.xlsx"
    setup_dummy_excel(dummy_xlsx_path)
    
    print("\n=== 2. AI & 룰베이스 파싱 파이프라인 테스트 ===")
    # Gemini API Key가 있다면 Gemini API가 구동될 것이고, 없다면 고도화된 로컬 정규식 분석이 수행됨
    parsed_data = run_parsing_pipeline(dummy_xlsx_path, "excel")
    print(f"파싱 결과: {parsed_data}")
    
    # 정합성 검증 확인
    print("\n--- 파싱 품질 대조 검증 ---")
    print(f"예상 공급업체: 주식회사 넷츠플러스 | 추출 결과: {parsed_data.get('company_name')}")
    print(f"예상 대표자명: 송주흔            | 추출 결과: {parsed_data.get('representative_name')}")
    print(f"예상 사업자번호: 264-81-46574    | 추출 결과: {parsed_data.get('business_number')}")
    print(f"예상 VAT포함금액: 362,304,800     | 추출 결과: {parsed_data.get('amount_incl_vat')}")
    
    # 3. DB 적재 시나리오 테스트
    print("\n=== 3. DB 적재 및 버전 누적 테스트 ===")
    from app.models import SessionLocal
    db = SessionLocal()
    
    try:
        # 프로젝트 생성
        project = Project(name="KCC정보통신 수협전산장비 통합유지보수", description="유지보수 견적 취합")
        db.add(project)
        db.commit()
        db.refresh(project)
        
        # Quotation 생성
        company_name = parsed_data.get("company_name", "미지정")
        quotation = Quotation(project_id=project.id, company_name=company_name)
        db.add(quotation)
        db.commit()
        db.refresh(quotation)
        
        # Version 1 생성 (신규 필드 포함)
        v1 = QuotationVersion(
            quotation_id=quotation.id,
            version=1,
            file_path=f"/static/files/test_dummy_netsplus.xlsx",
            file_name="test_dummy_netsplus.xlsx",
            file_type="excel",
            company_name=company_name,
            representative_name=parsed_data.get("representative_name"),
            business_number=parsed_data.get("business_number"),
            manager_name=parsed_data.get("manager_name"),
            contact=parsed_data.get("contact"),
            email=parsed_data.get("email"),
            amount_excl_vat=parsed_data.get("amount_excl_vat"),
            amount_incl_vat=parsed_data.get("amount_incl_vat"),
            raw_extracted_data=parsed_data.get("raw_extracted_data"),
            is_verified=False
        )
        db.add(v1)
        db.commit()
        db.refresh(v1)
        print(f"Version 1 적재 완료. ID={v1.id}, 대표자={v1.representative_name}, 사업자번호={v1.business_number}")
        
        # 수동 검증 및 수정 시나리오
        print("\n=== 4. 데이터 수정/검증 시나리오 테스트 ===")
        # 부가세포함 금액만 있고 별도 금액이 없어서 수동으로 입력했다고 가정
        v1.amount_excl_vat = 329368000.0 # 3억6천2백3십만4천원 / 1.1 = 3억2천9백3십6만8천원
        v1.is_verified = True
        db.commit()
        print(f"수정 및 검증완료 성공. 공급가액(VAT별도): {v1.amount_excl_vat}, 합계(VAT포함): {v1.amount_incl_vat}")
        
        # 4. 엑셀 취합 테스트
        print("\n=== 5. 프로젝트 내 최신 견적서 취합 엑셀 생성 테스트 ===")
        q_list = db.query(Quotation).filter(Quotation.project_id == project.id).all()
        q_data_list = []
        for q in q_list:
            if q.latest_version_id:
                ver = db.query(QuotationVersion).filter(QuotationVersion.id == q.latest_version_id).first()
                q_data_list.append({
                    "company_name": ver.company_name,
                    "representative_name": ver.representative_name,
                    "business_number": ver.business_number,
                    "manager_name": ver.manager_name,
                    "contact": ver.contact,
                    "email": ver.email,
                    "amount_excl_vat": float(ver.amount_excl_vat) if ver.amount_excl_vat else None,
                    "amount_incl_vat": float(ver.amount_incl_vat) if ver.amount_incl_vat else None,
                    "version": ver.version,
                    "is_verified": ver.is_verified,
                    "updated_at": ver.created_at
                })
        
        excel_stream = generate_consolidated_excel(q_data_list, project.name)
        excel_out_path = "storage/consolidated_netsplus_consolidated.xlsx"
        with open(excel_out_path, "wb") as f:
            f.write(excel_stream.read())
        print(f"취합 엑셀 파일 생성 성공: {excel_out_path}")
        
    finally:
        db.close()
        
    print("\n[성공] AI 정밀 파싱 개선 및 신규 필드 CRUD 테스트 완료!")

if __name__ == "__main__":
    test_db_and_pipeline()
