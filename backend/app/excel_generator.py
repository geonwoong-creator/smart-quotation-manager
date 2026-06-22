import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
from typing import List, Dict, Any

def generate_consolidated_excel(quotation_data: List[Dict[str, Any]], project_name: str) -> io.BytesIO:
    """프로젝트 내 최신 견적서 데이터를 취합하여 스타일링된 엑셀 바이너리를 생성합니다."""
    
    # 1. 데이터 프레임 생성
    columns_map = {
        "company_name": "업체명",
        "representative_name": "대표자명",
        "business_number": "사업자등록번호",
        "manager_name": "담당자명",
        "contact": "연락처",
        "email": "이메일",
        "amount_excl_vat": "견적금액 (VAT별도)",
        "amount_incl_vat": "견적금액 (VAT포함)",
        "version": "버전",
        "is_verified": "데이터 검증여부",
        "updated_at": "최종 업데이트 시간"
    }
    
    # 데이터 매핑
    rows = []
    for item in quotation_data:
        row = {}
        for key, col_name in columns_map.items():
            val = item.get(key)
            # 포맷 다듬기
            if key == "version" and val is not None:
                row[col_name] = f"V{val}"
            elif key == "is_verified":
                row[col_name] = "완료" if val else "대기(미검증)"
            elif key == "updated_at" and val:
                # 문자열 포맷팅 (datetime 객체 또는 string)
                if hasattr(val, "strftime"):
                    row[col_name] = val.strftime("%Y-%m-%d %H:%M")
                else:
                    row[col_name] = str(val)[:16]
            else:
                row[col_name] = val
        rows.append(row)
        
    df = pd.DataFrame(rows, columns=list(columns_map.values()))
    
    # 2. 엑셀 워크북 생성 및 스타일 적용
    output = io.BytesIO()
    
    # pandas excel writer (openpyxl engine)
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name="견적서 취합본", index=False)
        
        workbook = writer.book
        worksheet = writer.sheets["견적서 취합본"]
        
        # Gridlines 보이기
        worksheet.views.sheetView[0].showGridLines = True
        
        # 스타일 선언
        font_family = "맑은 고딕"
        header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid") # 차분한 스틸 블루
        
        body_font = Font(name=font_family, size=10, bold=False)
        align_center = Alignment(horizontal="center", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")
        
        # 테두리 설정
        thin_side = Side(border_style="thin", color="D9D9D9")
        border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        
        # 1행 헤더 스타일 설정
        worksheet.row_dimensions[1].height = 28
        for col_idx in range(1, len(columns_map) + 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center
            cell.border = border_all
            
        # 데이터 영역 스타일 및 정렬 설정
        for row_idx in range(2, len(rows) + 2):
            worksheet.row_dimensions[row_idx].height = 22
            for col_idx, col_key in enumerate(columns_map.keys(), start=1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.font = body_font
                cell.border = border_all
                
                # 정렬 및 서식
                if col_key in ["company_name", "email"]:
                    cell.alignment = align_left
                elif col_key in ["representative_name", "business_number", "manager_name", "contact", "version", "is_verified", "updated_at"]:
                    cell.alignment = align_center
                elif col_key in ["amount_excl_vat", "amount_incl_vat"]:
                    cell.alignment = align_right
                    cell.number_format = '#,##0' # 천단위 쉼표 포맷팅
        
        # 열 너비 자동 최적화
        for col in worksheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                # 한글 글자수 반영 보정 (한글은 2글자로 계상)
                val_len = sum(2 if ord(char) > 256 else 1 for char in val_str)
                if val_len > max_len:
                    max_len = val_len
            # 최소 너비 12, 최대 너비 40 보장
            worksheet.column_dimensions[col_letter].width = max(min(max_len + 3, 40), 12)
            
    output.seek(0)
    return output
