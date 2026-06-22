import os
import re
import json
import openpyxl
import pdfplumber
from typing import Dict, Any, Optional

# Gemini SDK 로드
try:
    import google.generativeai as genai
    HAS_GEMINI_SDK = True
except ImportError:
    HAS_GEMINI_SDK = False

def load_env():
    # 프로젝트 루트 또는 상위 디렉토리의 .env 로드
    paths = [
        os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", ".env")),
        os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".env")),
        os.path.abspath(".env")
    ]
    for p in paths:
        if os.path.exists(p):
            try:
                with open(p, "r", encoding="utf-8-sig") as f:
                    for line in f:
                        line = line.strip()
                        if line and not line.startswith("#") and "=" in line:
                            k, v = line.split("=", 1)
                            key_name = k.strip()
                            val_value = v.strip().strip('"').strip("'")
                            os.environ[key_name] = val_value
                            if key_name == "GEMINI_API_KEY":
                                if val_value and val_value != "YOUR_GEMINI_API_KEY_HERE":
                                    masked = val_value[:6] + "..." + val_value[-4:] if len(val_value) > 10 else "단문 키"
                                    print(f"[System] .env 로드 성공: GEMINI_API_KEY={masked} ({p})")
                                else:
                                    print(f"[System] 템플릿 기본 키 감지: 키를 교체해 주세요. ({p})")
                # .env 파일을 성공적으로 로드했으므로 루프 탈출
                break
            except Exception as e:
                print(f"[System] .env 로드 오류: {e}")


load_env()


GEMINI_API_KEY = os.environ.get("GEMINI_API_KEY", "")


def clean_amount(val_str: Optional[str]) -> Optional[float]:
    """금액 텍스트에서 쉼표, 한글 통화 단위, 괄호 등을 제거하고 float로 변환합니다."""
    if not val_str:
        return None
    try:
        # 숫자만 남김
        cleaned = re.sub(r'[^\d]', '', val_str)
        return float(cleaned) if cleaned else None
    except ValueError:
        return None

def extract_email_phone(text: str) -> tuple[Optional[str], Optional[str]]:
    """텍스트에서 이메일과 연락처를 추출합니다."""
    email_pattern = r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}'
    phone_pattern = r'(?:010|02|031|032|042|051|052|053|054|055|061|062|064)-\d{3,4}-\d{4}'
    
    email_match = re.search(email_pattern, text)
    phone_match = re.search(phone_pattern, text)
    
    email = email_match.group(0) if email_match else None
    phone = phone_match.group(0) if phone_match else None
    return email, phone

def extract_text_from_pdf(file_path: str) -> str:
    """PDF 파일 내의 선택 가능한 텍스트를 추출합니다."""
    text = ""
    try:
        with pdfplumber.open(file_path) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"
    except Exception as e:
        print(f"PDF 텍스트 추출 중 오류: {e}")
    return text

def parse_old_xls_text_dump(file_path: str) -> str:
    """구형 .xls 파일의 텍스트 데이터를 xlrd 라이브러리를 통해 추출합니다."""
    texts = []
    try:
        import xlrd
        wb = xlrd.open_workbook(file_path)
        for sheet in wb.sheets():
            sheet_title_logged = False
            # 최대 200행, 30열까지 안전 스캔
            max_r = min(sheet.nrows, 200)
            max_c = min(sheet.ncols, 30)
            for r in range(max_r):
                row_texts = []
                for c in range(max_c):
                    val = sheet.cell_value(r, c)
                    if val is not None and str(val).strip() != "":
                        row_texts.append(str(val).strip())
                if row_texts:
                    if not sheet_title_logged:
                        texts.append(f"=== Sheet: {sheet.name} ===")
                        sheet_title_logged = True
                    texts.append(" | ".join(row_texts))
    except Exception as e:
        print(f"구형 .xls 파일 파싱 중 오류: {e}")
    return "\n".join(texts)

def parse_excel_text_dump(file_path: str) -> str:
    """Excel(.xlsx, .xls)의 전체 워크시트에서 셀 텍스트를 추출하여 LLM용 평면 텍스트로 만듭니다."""
    
    # 구형 .xls 포맷 감지 시 xlrd 파서로 라우팅
    if file_path.lower().endswith(".xls"):
        print("[System] 구형 .xls 파일로 감지되어 xlrd 파서를 작동합니다.")
        return parse_old_xls_text_dump(file_path)

    def extract_from_workbook(wb) -> str:
        local_texts = []
        for sheet in wb.worksheets:
            sheet_title_logged = False
            # openpyxl max_row/max_column 오작동 대비 최소 100행, 25열 스캔 보장
            det_r = sheet.max_row if sheet.max_row else 100
            max_r = min(max(det_r + 1, 100), 200)
            det_c = sheet.max_column if sheet.max_column else 25
            max_c = min(max(det_c + 1, 25), 40)
            for r in range(1, max_r):
                row_texts = []
                for c in range(1, max_c):
                    val = sheet.cell(row=r, column=c).value
                    if val is not None:
                        row_texts.append(str(val).strip())
                if row_texts:
                    if not sheet_title_logged:
                        local_texts.append(f"=== Sheet: {sheet.title} ===")
                        sheet_title_logged = True
                    local_texts.append(" | ".join(row_texts))
        return "\n".join(local_texts)

    # .xlsx drawings/NULL 에러 우회를 위한 read_only=True 옵션 적용 하이브리드 로드
    try:
        # 1차 시도: data_only=True, read_only=True (drawings 분석을 스킵하여 에러 우회 및 수식 결과 덤프)
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        text_content = extract_from_workbook(wb)
        wb.close()
        
        # 만약 긁어온 데이터가 거의 없다면 (수식 계산 캐시가 없는 엑셀인 경우)
        if len(text_content.strip()) < 30:
            print("[Warning] read_only 결과값 추출 실패. 수식 포함(read_only=True, data_only=False) 로드합니다.")
            wb_raw = openpyxl.load_workbook(file_path, data_only=False, read_only=True)
            text_content = extract_from_workbook(wb_raw)
            wb_raw.close()
            
        return text_content
    except Exception as e:
        print(f"Excel 텍스트 덤프 중 오류 발생 (read_only=True): {e}")
        # 오류 발생 시 read_only 해제하고 일반 복구 로드로 시도
        try:
            wb_raw = openpyxl.load_workbook(file_path, data_only=False)
            text_content = extract_from_workbook(wb_raw)
            wb_raw.close()
            return text_content
        except Exception as e2:
            print(f"Excel 백업 일반 덤프마저 실패: {e2}")
    return ""



# -----------------
# 1. Gemini LLM 파서 모듈 (핵심)
# -----------------
def parse_text_with_gemini(text: str) -> Optional[Dict[str, Any]]:
    """Gemini API를 사용하여 견적서 텍스트로부터 정밀한 구조화 데이터를 추출합니다."""
    if not HAS_GEMINI_SDK or not GEMINI_API_KEY:
        print("Gemini API Key 혹은 SDK가 설정되어 있지 않아 룰베이스 파서로 우회합니다.")
        return None
        
    try:
        # API 설정
        genai.configure(api_key=GEMINI_API_KEY)
        
        # 구조화된 출력을 위한 스키마 및 프롬프트 작성
        prompt = f"""
당신은 기업 견적서 및 세금 서류 분석 전문 AI입니다. 
아래 제공된 견적서 텍스트에서 지정된 필드 데이터를 정확히 분석하여 JSON 형식으로만 추출해 주세요.

[요청 필드 정의]
1. "company_name": 견적을 발송한 공급자(사업자)의 상호명 (예: "주식회사 넷츠플러스"). 
   ★주의★: 'KCC정보통신귀중'처럼 '귀하', '귀중', '수신'으로 수신처 명기된 회사는 견적을 "받는" 곳이므로 절대 company_name으로 쓰면 안 됩니다. 공급자의 상호를 넣어야 합니다.
2. "representative_name": 공급자 측의 대표자 성명 (예: "송주흔"). '대표자:', '대표:', '대표자명:' 뒤에 명시된 이름입니다.
3. "business_number": 공급자 측의 사업자등록번호 (예: "264-81-46574"). 세자리-두자리-다섯자리 구조의 번호입니다.
4. "manager_name": 공급자 측 실무 담당자 성명 (예: "홍길동 대리"). 대표자명과 실무 담당자명이 다를 경우 구분해 주세요.
5. "contact": 연락처 (전화번호 혹은 휴대폰번호).
6. "email": 이메일 주소.
7. "amount_excl_vat": 부가세(VAT) 별도 금액 (순수 공급가액). 숫자만 추출해야 합니다. 
   ★주의★: 원본에 '부가세포함' 금액만 명시된 경우, 이를 1.1로 나누어 계산하지 말고, 해당 금액은 amount_incl_vat에 넣고 amount_excl_vat는 그에 맞는 부가세별도 공급가액(수식상 공급가액)을 찾아 매핑해야 합니다.
8. "amount_incl_vat": 부가세(VAT) 포함 금액 (합계금액). '총합계', '총공급금액(부가세포함)' 등으로 표현됩니다. 

[견적서 텍스트]
\"\"\"
{text}
\"\"\"

반환할 JSON 스키마 형식:
{{
  "company_name": string | null,
  "representative_name": string | null,
  "business_number": string | null,
  "manager_name": string | null,
  "contact": string | null,
  "email": string | null,
  "amount_excl_vat": number | null,
  "amount_incl_vat": number | null
}}
"""
        # 최신 가성비 모델인 gemini-3.5-flash 모델 호출
        model = genai.GenerativeModel("gemini-3.5-flash")
        response = model.generate_content(
            prompt,
            generation_config={"response_mime_type": "application/json"}
        )
        
        # 결과 JSON 파싱
        if response.text:
            result = json.loads(response.text.strip())
            # 기본 검증 및 보완
            if not result.get("company_name"):
                result["company_name"] = "미지정 업체"
            return result
            
    except Exception as e:
        print(f"Gemini API 호출 중 실패: {e}")
        return None
        
    return None

# -----------------
# 2. 로컬 룰베이스 파서 (Fallback 백업용)
# -----------------
def parse_text_data_local_regex(text: str) -> Dict[str, Any]:
    """Gemini API가 동작하지 않을 때 작동하는 개선된 로컬 정규식 파서입니다."""
    data = {
        "company_name": None,
        "representative_name": None,
        "business_number": None,
        "manager_name": None,
        "contact": None,
        "email": None,
        "amount_excl_vat": None,
        "amount_incl_vat": None,
        "raw_extracted_data": {
            "parser_type": "local_regex_fallback",
            "warning": "Gemini API Key가 등록되지 않았거나 오류가 발생해 로컬 파서로 우회 분석되었습니다."
        }
    }
    
    # 1. 사업자번호 추출
    biz_match = re.search(r'(?:등록번호|사업자번호|사업자등록번호)\s*[:：]?\s*(\d{3}-\d{2}-\d{5})', text)
    if biz_match:
        data["business_number"] = biz_match.group(1).strip()
        
    # 2. 대표자명 추출
    rep_match = re.search(r'(?:대표자|대표|대표자명)\s*[:：]?\s*([^\s\n\r(]+)', text)
    if rep_match:
        data["representative_name"] = rep_match.group(1).strip()

    # 3. 공급사 상호명 추출 (귀중/수신 단어 필터링 우회 적용)
    company_name = None
    # '상 호' 또는 '상호'로 명시된 줄 매칭
    co_match = re.search(r'(?:상\s*호|공급자|업체명)\s*[:：]\s*([^\n\r|]+)', text)
    if co_match:
        val = co_match.group(1).strip()
        if "귀중" not in val and "귀하" not in val:
            company_name = val
            
    # 매칭 실패 시 텍스트 전체에서 상호 매칭 시도
    if not company_name:
        lines = text.split("\n")
        for line in lines:
            if "상호" in line or "공급자" in line:
                m = re.search(r'(?:상호|공급자)\s*[:：]?\s*([^\n\r|]+)', line)
                if m:
                    val = m.group(1).strip()
                    if "귀중" not in val and "귀하" not in val:
                        company_name = val
                        break
                        
    data["company_name"] = company_name if company_name else "미지정 업체"

    # 4. 담당자명
    manager_match = re.search(r'담당자\s*[:：]\s*([^\s\n\r|]+)', text)
    if manager_match:
        val = manager_match.group(1).strip()
        # 귀중 회사 담당자와 공급사 담당자 구분 시도
        if "귀중" not in text.split(val)[0]: # 단순 위치 휴리스틱
            data["manager_name"] = val
            
    # 5. 연락처 및 이메일
    email, contact = extract_email_phone(text)
    data["email"] = email
    data["contact"] = contact

    # 6. 금액 추출 고도화 (부가세포함과 별도 구분 필터링)
    # 총공급금액, 합계 등 부가세포함 금액 키워드
    incl_match = re.search(r'(?:합계금액|합계|총공급금액|총액|총금액)\s*[:：]?\s*([\d,]+)\s*(?:원)?', text)
    # 공급가액, 공급가 등 부가세별도 금액 키워드
    excl_match = re.search(r'(?:공급가액|공급가|소계)\s*[:：]?\s*([\d,]+)\s*(?:원)?', text)
    
    amount_incl = clean_amount(incl_match.group(1)) if incl_match else None
    amount_excl = clean_amount(excl_match.group(1)) if excl_match else None
    
    # 텍스트에 "부가세포함" 또는 "VAT포함" 이 명시되어 있는데 excl에 그 금액이 들어가 있다면 스왑 처리
    if "부가세포함" in text or "VAT포함" in text:
        if amount_excl and not amount_incl:
            amount_incl = amount_excl
            amount_excl = None
            
    data["amount_incl_vat"] = amount_incl
    data["amount_excl_vat"] = amount_excl

    return data

# -----------------
# 3. 메인 통합 파이프라인
# -----------------
def run_parsing_pipeline(file_path: str, file_type: str) -> Dict[str, Any]:
    """파일 타입에 맞추어 텍스트를 추출하고, Gemini LLM 파서 혹은 로컬 룰베이스 파서를 구동합니다."""
    text_content = ""
    
    # 1) 파일에서 텍스트 수집
    if file_type == "excel":
        text_content = parse_excel_text_dump(file_path)
    elif file_type == "pdf":
        text_content = extract_text_from_pdf(file_path)
        # 만약 스캔본 PDF라 텍스트 추출이 불가능하면, OCR 또는 LLM Vision 구동 준비
        # 여기서는 기본적으로 텍스트를 추출한 것으로 가정하고, 없다면 빈 텍스트 상태로 LLM에 넘깁니다.
    elif file_type in ["image", "png", "jpg", "jpeg"]:
        # 이미지의 경우 실무적으로는 Vision API 연동
        # 여기서는 텍스트가 없는 스캔 이미지를 위해 더미 텍스트를 할당하거나 Vision 모드를 활용
        text_content = "[스캔 이미지 파일 분석 요청]\n(실무 적용 시 Gemini 1.5 Pro / Flash Multi-modal 호출을 통해 이미지를 직접 분석합니다.)"
        
    # 2) 파싱 알고리즘 가동
    # 2-1) 우선 Gemini API 시도 (Key 가 있고 SDK가 로드된 경우)
    if GEMINI_API_KEY:
        print("정밀 데이터 분석을 위해 Gemini API 파서를 호출합니다.")
        llm_result = parse_text_with_gemini(text_content)
        if llm_result:
            # 성공 시 메타데이터 추가 후 반환
            llm_result["raw_extracted_data"] = {
                "parser_type": "gemini_api_llm",
                "model": "gemini-1.5-flash",
                "extracted_fields_count": sum(1 for v in llm_result.values() if v is not None),
                "text_snippet": text_content[:300]
            }
            return llm_result
            
    # 2-2) 실패 혹은 API Key 미등록 시 로컬 정규식 파서(Fallback) 실행
    print("로컬 룰베이스 파서를 작동시킵니다.")
    local_result = parse_text_data_local_regex(text_content)
    
    # 디버그용으로 덤프 텍스트 일부 제공
    if local_result.get("raw_extracted_data"):
        local_result["raw_extracted_data"]["text_snippet"] = text_content[:300]
        
    return local_result
